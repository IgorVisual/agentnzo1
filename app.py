import os
import pandas as pd
import streamlit as st
import random
from openpyxl import Workbook

# Функция для инициализации таблицы-шаблона с заголовками и подзаголовками
def initialize_template_table(headers, subheaders):
    template_df = pd.DataFrame({
        "Заголовки": headers,
        "Подзаголовки": subheaders,
        "Выбранные товары": [[] for _ in range(len(headers))]  # Пустые списки для товаров
    })
    return template_df

# Функция для маппинга данных в итоговую таблицу и сохранения в Excel
def save_to_excel(mapped_data, output_filename):
    wb = Workbook()
    ws = wb.active

    # Записываем данные в Excel по группам
    row_num = 1

    for category, items in mapped_data.items():
        # Вставляем название категории
        ws[f'A{row_num}'] = category
        row_num += 1

        # Вставляем заголовки
        ws[f'A{row_num}'] = "Наименование"
        ws[f'B{row_num}'] = "Цена, руб"
        ws[f'C{row_num}'] = "Количество"
        row_num += 1

        # Вставляем данные
        for item in items:
            ws[f'A{row_num}'] = item["Наименование"]
            ws[f'B{row_num}'] = item.get("Цена", "")  # Цена может быть не указана
            ws[f'C{row_num}'] = item.get("Количество", "")  # Количество может быть не указано
            row_num += 1

        # Пустая строка между категориями
        row_num += 1

    # Сохранение файла
    wb.save(output_filename)
    print(f"Файл успешно сохранен как {output_filename}")

# Загрузка данных для второй таблицы (например, текущий файл Excel)
file_path = 'Каталог_Чинт.xlsx'  # Укажи путь к файлу Excel
df2 = pd.read_excel(file_path, header=14)  # Чтение с заголовками из 15-й строки

# Очищаем заголовки второй таблицы от лишних пробелов и спецсимволов
df2.columns = df2.columns.str.strip()
df2.columns = df2.columns.str.replace(r'[\n\r\t]', ' ', regex=True)
df2.columns = df2.columns.str.replace(r'\s+', ' ', regex=True)

# Фильтруем корректные столбцы
valid_columns = [col for col in df2.columns if not col.startswith("Unnamed")]
df2 = df2[valid_columns]

# Столбцы по умолчанию
default_columns = [col for col in ["Наименование", "Тариф с НДС, руб"] if col in valid_columns]

# Установка кастомных стилей для центрирования и уменьшения отступов
st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1.7rem;
        padding-bottom: 0rem;
        padding-left: 1rem;
        padding-right: 1rem;
        text-align: center;  /* Центрирование текста */
    }
    .css-1p05t01 {
        padding: 0;
    }
    .st-dataframe {
        width: 100%;
    }
    .stSelectbox label, .stTextInput label {
        text-align: center;  /* Центрирование текста в полях ввода */
    }
    .stSelectbox div, .stTextInput div {
        margin-left: auto;
        margin-right: auto;
        text-align: center;
    }
    .stCheckbox div {
        margin-left: auto;
        margin-right: auto;
        text-align: center;  /* Центрирование чекбоксов */
    }
    .checkbox-container {
        display: flex;
        align-items: center;
        justify-content: space-between;
        margin-bottom: 0.5rem;
    }
    </style>
    """, 
    unsafe_allow_html=True
)

# Интерфейс Streamlit
st.title("Агент NZO")

# Фиксированные переменные для раздела
categories = ["Корпус", "Отсек высоковольтного выключателя", "Отсек РЗА", "Прочее"]

# Инициализация таблицы-шаблона с фиксированными категориями и подкатегорией "Оборудование"
template_df = initialize_template_table(categories, ["Оборудование"] * len(categories))

# Переменная для сохранения выбранных товаров
if 'selected_items' not in st.session_state:
    st.session_state.selected_items = {category: [] for category in categories}

# Переменная для сохранения текущего состояния строк с выбором товаров
if 'current_selected_rows' not in st.session_state:
    st.session_state.current_selected_rows = []

# Второй блок: таблица с файла из Каталог_Чинт.xlsx с чекбоксами
st.subheader("Поиск оборудования")

# Добавление поля для поиска и фильтров на одной строке
with st.container():
    col1, col2 = st.columns([3, 1])
    with col1:
        search_query = st.text_input("Поиск товаров", "", key="search_query")
    with col2:
        selected_header = st.selectbox("Раздел", categories)

# Поиск по таблице
if search_query.strip():
    filtered_df2 = df2[df2.apply(lambda row: row.astype(str).str.contains(search_query, case=False).any(), axis=1)]
    show_table = True
else:
    filtered_df2 = pd.DataFrame(columns=df2.columns)  # Пустая таблица
    show_table = False

# Используем кастомные блоки для отображения товаров
if show_table:
    for idx, row in filtered_df2.iterrows():
        item_name = row["Наименование"]
        item_price = row.get("Тариф с НДС, руб", "Цена не указана")
        
        # Оформляем блок с товаром и ценой
        with st.container():
            col1, col2 = st.columns([4, 1])
            with col1:
                st.markdown(f"**{item_name}**")
                st.markdown(f"*Цена*: {item_price} руб")
            with col2:
                # Добавляем индекс строки (idx) к ключу для уникальности
                selected = st.checkbox("Выбрать", key=f"{item_name}_{selected_header}_{idx}")
                # Сохраняем выбор в текущем разделе
                if selected and item_name not in st.session_state.current_selected_rows:
                    st.session_state.current_selected_rows.append(item_name)
                    st.session_state.selected_items[selected_header].append(item_name)

# Третий блок: таблица-шаблон для выбранных товаров на всю ширину
st.subheader("Итоговый файл для просчета")

# Отображаем таблицу с заголовками и выбранными товарами
selected_df = pd.DataFrame({
    "Заголовки": categories,
    "Выбранные товары": [", ".join(st.session_state.selected_items[header]) for header in categories]
})
st.dataframe(selected_df, use_container_width=True, hide_index=True)

# Четвертый блок: Сохранение в Excel
if st.button("Сохранить в Excel"):
    # Маппинг выбранных данных в итоговую таблицу
    mapped_data = {}
    for category, items in st.session_state.selected_items.items():
        if items:
            mapped_data[category] = [{"Наименование": item} for item in items]

    # Сохранение в Excel файл
    save_to_excel(mapped_data, 'mapped_data.xlsx')
    
    # Сброс выбранных товаров после сохранения
    st.session_state.selected_items = {category: [] for category in categories}
    st.session_state.current_selected_rows = []
    
    # Создание ссылки для скачивания файла
    with open("mapped_data.xlsx", "rb") as file:
        st.download_button(
            label="Скачать Excel файл",
            data=file,
            file_name="mapped_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    st.success("Файл сохранен как mapped_data.xlsx и доступен для скачивания!")
